import sys
sys.path.insert(0, 'E:/CS/Graduation Project/-A-learning-based-semi-Humanoid-Robot-for-On-campus-Surveillance-main/VoiceCommands')
import Project_VoiceCommands as vc

# Python program to read an excel file
# import openpyxl module
import openpyxl
def ReadFileData_Day (Path , Day ):
    
    
    
    # Give the location of the file
    try :
        path = "StudentsData/" +Path+".xlsx"
    
    except (FileNotFoundError, IOError):
        print("Wrong file or file path")
    
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    Rows=sheet_obj.max_row
    Columns=sheet_obj.max_column
    #DayList=[]
    CourseTiming=[]
    CourseName=[]
    Starting=[]
    Ending=[]
    FullTime=""
    SecondPart =""
    Room=[]
    #------------------ 1st For Loop To Find the day -------------#
    for i in range (2,Columns+1):
        Day_Sch = sheet_obj.cell(row = 1, column = i)
        
       
        #---------Checking if the day is the same day User defined------#
        if Day_Sch.value == Day:
            #------------------ 2nd For Loop To Find the Course -------------#
            for j in range(2,Rows+1):
                cell_obj = sheet_obj.cell(row = j, column = i)
                cell_course_name = sheet_obj.cell(row = j, column = 1)
                CourseTiming.append(cell_obj.value)
                CourseName.append(cell_course_name.value)
                
        
        
    
    for i in range(len(CourseName)):
        print (CourseName[i])
        
        Timing = CourseTiming[i]
        
        if Timing is None :
            FullTime = "None"
            Starting.append(FullTime)
            Ending.append(FullTime)
            Room.append(FullTime)
            print ("Start In "+Starting[i])
            print ("End In "+Ending[i])
            print ("Room is "+Room[i])
        else :
    
            FullTime = Timing.split("-")
            Starting.append(FullTime[0])
            SecondPart = FullTime[1].split("M")
            Ending.append(SecondPart[0]+"M")
            Room.append(SecondPart[1])
            print ("Start In "+Starting[i])
            print ("End In "+Ending[i])
            print ("Room is "+Room[i])
            
        
        
def ReadFileData_Time (Path , Day, Time ):
    
    
    # Give the location of the file
    try :
        path = "E:/CS/Graduation Project/-A-learning-based-semi-Humanoid-Robot-for-On-campus-Surveillance-main/StudentsData/" +Path+".xlsx"
    
    except (FileNotFoundError, IOError):
        print("Wrong file or file path")
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    Rows=sheet_obj.max_row
    Columns=sheet_obj.max_column
    #DayList=[]
    CourseTiming=[]
    CourseName=[]
    Starting=[]
    Ending=[]
    FullTime=""
    SecondPart =""
    Room=[]
    Flag= False
    time=int(Time)
    #------------------ 1st For Loop To Find the day -------------#
    for i in range (2,Columns+1):
        Day_Sch = sheet_obj.cell(row = 1, column = i)
        
       
        #---------Checking if the day is the same day User defined------#
        if Day_Sch.value == Day:
            #------------------ 2nd For Loop To Find the Course -------------#
            for j in range(2,Rows+1):
                cell_obj = sheet_obj.cell(row = j, column = i)
                cell_course_name = sheet_obj.cell(row = j, column = 1)
                CourseTiming.append(cell_obj.value)
                CourseName.append(cell_course_name.value)
                
        
        
    
    for i in range(len(CourseName)):
        #print (CourseName[i])
        
        Timing = CourseTiming[i]
        
        if Timing is None :
            FullTime = "None"
            Starting.append(FullTime)
            Ending.append(FullTime)
            Room.append(FullTime)
            #print ("Start In "+Starting[i])
            #print ("End In "+Ending[i])
            
        else :
    
            FullTime = Timing.split("-")
            Starting.append(FullTime[0])
            SecondPart = FullTime[1].split("M")
            Ending.append(SecondPart[0]+"M")
            Room.append(SecondPart[1])
            #print ("Start In "+Starting[i])
           # print ("End In "+Ending[i])
    
        
    for i in range (len(CourseName)):
        #if Starting[i]>= 12
        if Starting[i] == "None":
            break
        
        else :
            
            FullStart = Starting[i]
            FullEnd = Ending[i]
            
            PartialStart=FullStart.split(" ")
            PartialEnd=FullEnd.split(" ")
            
            NearStart=PartialStart[0]
            NearEnd = PartialEnd[0]
            
            Start=NearStart.split(":")
            End=NearEnd.split(":")
            
            
            first=int(Start[0])
            last = int(End[0])
            if first <= time <= last :
                Flag = True
            if Flag == True :
                print("You are Busy At This Time Having :-")
                print(CourseName[i])
                print ("Start In "+Starting[i])
                print ("End In "+Ending[i])
                print("Room"+Room[i])
                vc.speak("You are Busy At This Time Having :-")
                vc.speak(CourseName[i])
                vc.speak ("Starts In "+Starting[i])
                vc.speak ("Ends In "+Ending[i])
                vc.speak("Room"+Room[i])
                break
            
            
    if Flag==False: 
        print("You don't Have Lectures at this time ")
        vc.speak("You don't Have Lectures at this time ")
            
    
        

if __name__ == '__main__':
    Name="Pierre Malak"
    day="Sunday"
    time ="17"
    #ReadFileData_Day(Name ,day)
      
    ReadFileData_Time(Name , day, time)
